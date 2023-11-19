# standard library
from __future__ import annotations
import os
from pathlib import Path
from typing import List, Dict, Tuple

# pandas
import pandas as pd

# openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# wsi
from core import constants
from core import utils

# PIL
from PIL import Image

# joblib
from joblib import delayed, dump, load

# openslide
if hasattr(os, 'add_dll_directory'):
    # Python >= 3.8 on Windows
    with os.add_dll_directory(constants.openslide_path):
        import openslide
else:
    import openslide


dataset_name_column = 'DatasetName'
batch_id_column = 'BatchID'
slide_name_column = 'SlideName'
slide_id_column = 'SlideID'
block_id_column = 'BlockID'
tissue_id_column = 'TissueID'
sample_id_column = 'SampleID'


class SlideData:
    def __init__(self, path: Path):
        self._path = path

    @property
    def slide_name(self) -> str:
        return self._path.stem

    @property
    def slide_id(self) -> str:
        return self.slide_name.replace('_', '/')

    @property
    def block_id(self) -> str:
        return self.slide_id.rsplit('/', 1)[0]

    @property
    def tissue_id(self) -> str:
        return self.block_id.rsplit('/', 1)[0]

    @property
    def sample_id(self) -> str:
        return self.tissue_id.rsplit('/', 1)[0]

    @property
    def sample_id(self) -> str:
        return self.tissue_id.rsplit('/', 1)[0]

    @property
    def batch_id(self) -> str:
        return self._path.parts[-2].capitalize()

    @property
    def dataset_name(self) -> str:
        batch_id = self.batch_id
        if batch_id.startswith('Carmel'):
            return 'Carmel'
        if batch_id.startswith('Her2'):
            return 'Her2'

    def get_row(self) -> List[str]:
        return [self.dataset_name, self.batch_id, self.slide_name, self.slide_id, self.block_id, self.tissue_id, self.sample_id]

    @staticmethod
    def get_column_names() -> List[str]:
        return [
            dataset_name_column,
            batch_id_column,
            slide_name_column,
            slide_id_column,
            block_id_column,
            tissue_id_column,
            sample_id_column
        ]


class SlidesMapping:
    # def __init__(self, paths: List[Path]):
    #     self._paths = paths
    #     self._slides_data: List[SlideData] = []
    #     self._block_id_to_slide_ids: Dict[str, List[str]] = {}
    #     self._slide_id_to_path: Dict[str, Path] = {}
    #     self._block_id_to_paths: Dict[str, List[Path]] = {}
    #
    #     for path in paths:
    #         print(path)
    #         slide_data = SlideData(path=path)
    #         slide_id = slide_data.slide_id
    #         block_id = slide_data.block_id
    #
    #         if block_id not in self._block_id_to_slide_ids:
    #             self._block_id_to_slide_ids[block_id] = []
    #
    #         self._block_id_to_slide_ids[block_id].append(slide_id)
    #         self._slide_id_to_path[slide_id] = path
    #
    #         if block_id not in self._block_id_to_paths:
    #             self._block_id_to_paths[block_id] = []
    #         self._block_id_to_paths[block_id].append(path)
    #         self._slides_data.append(slide_data)
    #
    #     self._df = self._create_dataframe()

    def __init__(self, df: pd.DataFrame):
        self._df = df

    @staticmethod
    def from_file_paths(paths: List[Path]) -> SlidesMapping:
        slide_paths = SlidesMapping._list_files(paths=paths, ext='mrxs')
        slides_data: List[SlideData] = []
        for slide_path in slide_paths:
            slide_data = SlideData(path=slide_path)
            slides_data.append(slide_data)

        data = [slide_data.get_row() for slide_data in slides_data]
        columns = SlideData.get_column_names()
        df = pd.DataFrame(data, columns=columns)
        return SlidesMapping(df=df)

    @staticmethod
    def from_excel(path: Path) -> SlidesMapping:
        df = pd.read_excel(io=path)
        return SlidesMapping(df=df)

    @staticmethod
    def _remove_non_digits(s: str) -> str:
        return ''.join(char for char in s if char.isdigit())

    @staticmethod
    def _list_files(paths: List[Path], ext: str) -> List[Path]:
        listed_file_paths = []
        for path in paths:
            for root, dirs, files in os.walk(path):
                for file in files:
                    if file.endswith(ext):
                        full_path = os.path.normpath(os.path.join(root, file))
                        listed_file_paths.append(Path(full_path))
                break
        return listed_file_paths

    @staticmethod
    def _save_thumbnails_for_block_id(block_id: str, size: Tuple[int, int], input_path: Path, output_path: Path, dump_path: Path):
        df = load(str(dump_path))
        slide_mapping = SlidesMapping(df=df)
        block_id_path = output_path / Path(block_id)
        block_id_path.mkdir(parents=True, exist_ok=True)
        slide_paths = slide_mapping.get_slide_paths_by_block_id(block_id=block_id, base_path=input_path)
        for slide_path in slide_paths:
            slide = openslide.OpenSlide(str(slide_path))
            thumbnail_image = slide.get_thumbnail(size=size)
            thumbnail_path = block_id_path / Path(f'{slide_path.stem}.png')
            print(f'Saving thumbnail: {thumbnail_path}')
            thumbnail_image.save(thumbnail_path)

    def get_slide_paths_by_block_id(self, block_id: str, base_path: Path) -> List[Path]:
        print(block_id_column)
        mask = self._df[block_id_column] == block_id
        filtered_df = self._df[mask]

        def row_to_path(row: pd.Series) -> Path:
            if row[dataset_name_column] == 'Carmel':
                batch_number = int(SlidesMapping._remove_non_digits(s=row[batch_id_column]))
                if 1 <= batch_number <= 8:
                    subfolder = '1-8'
                elif batch_number > 8:
                    subfolder = '9-11'
            elif row[dataset_name_column] == 'Her2':
                batch_number = int(row[batch_id_column].rsplit('_', 1)[1])
                subfolder = 'Her2'

            path = Path(f'{base_path}/{subfolder}/Batch_{batch_number}/{row[batch_id_column].upper()}/{row[slide_name_column]}.mrxs')
            return path

        slide_paths = filtered_df.apply(row_to_path, axis=1).tolist()
        return slide_paths

    def get_block_ids(self) -> List[str]:
        return self._df[block_id_column].unique().tolist()

    def save_thumbnails(self, input_path: Path, output_path: Path, size: Tuple[int, int]):
        block_ids = self.get_block_ids()
        array = self._df.values
        dump_path = output_path / 'dataframe.joblib'
        dump(value=array, filename=str(dump_path))
        utils.ProgressParallel(
            use_tqdm=False,
            n_jobs=1,
            total=len(block_ids))(
            delayed(self._save_thumbnails_for_block_id)(
                block_id=block_id,
                size=size,
                input_path=input_path,
                output_path=output_path,
                dump_path=dump_path) for block_id in block_ids)

    @property
    def block_id_to_slide_ids(self) -> Dict[str, List[str]]:
        return self._block_id_to_slide_ids

    @property
    def thumb_id_to_path(self) -> Dict[str, Path]:
        return self._slide_id_to_path

    @property
    def block_id_to_paths(self) -> Dict[str, List[Path]]:
        return self._block_id_to_paths

    @staticmethod
    def _get_block_id(thumb_id: str) -> str:
        return thumb_id.rsplit('_', 1)[0]

    def _create_dataframe(self) -> pd.DataFrame:
        data = [slide_data.get_row() for slide_data in self._slides_data]
        columns = SlideData.get_column_names()
        df = pd.DataFrame(data, columns=columns)
        return df

    def save_dataframe(self, path: Path):
        path_str = str(path)
        self._df.to_excel(path_str, index=False, engine='openpyxl')
        workbook = load_workbook(path_str)
        sheet = workbook.active

        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width

        workbook.save(path_str)

    def load_dataframe(self, path: Path):
        self._df = pd.read_excel(io=path)

    def get_block_ids_with_multiple_her2(self) -> List[str]:
        grouped_df = self._df.groupby(block_id_column)
        bla = grouped_df.apply(lambda x: (x[dataset_name_column] == 'Her2').sum())
        pass


if __name__ == '__main__':
    input_path = Path('/mnt/gipmed_new/Data/Breast/Carmel')
    output_path = Path('/mnt/gipmed_new/Data/Breast/Carmel/HE_IHC')
    # output_path = Path('C:\\out_test')

    base_paths = [
        Path('/mnt/gipmed_new/Data/Breast/Carmel/1-8/Batch_1/CARMEL1'),
        Path('/mnt/gipmed_new/Data/Breast/Carmel/1-8/Batch_2/CARMEL2'),
        Path('/mnt/gipmed_new/Data/Breast/Carmel/1-8/Batch_3/CARMEL3'),
        Path('/mnt/gipmed_new/Data/Breast/Carmel/1-8/Batch_4/CARMEL4'),
        Path('/mnt/gipmed_new/Data/Breast/Carmel/1-8/Batch_5/CARMEL5'),
        Path('/mnt/gipmed_new/Data/Breast/Carmel/1-8/Batch_6/CARMEL6'),
        Path('/mnt/gipmed_new/Data/Breast/Carmel/1-8/Batch_7/CARMEL7'),
        Path('/mnt/gipmed_new/Data/Breast/Carmel/1-8/Batch_8/CARMEL8'),
        Path('/mnt/gipmed_new/Data/Breast/Carmel/9-11/Batch_9/CARMEL9'),
        Path('/mnt/gipmed_new/Data/Breast/Carmel/9-11/Batch_10/CARMEL10'),
        Path('/mnt/gipmed_new/Data/Breast/Carmel/9-11/Batch_11/CARMEL11'),
        Path('/mnt/gipmed_new/Data/Breast/Carmel/Her2/Batch_1/Her2_1'),
        Path('/mnt/gipmed_new/Data/Breast/Carmel/Her2/Batch_2/HER2_2'),
        Path('/mnt/gipmed_new/Data/Breast/Carmel/Her2/Batch_3/HER2_3'),
        Path('/mnt/gipmed_new/Data/Breast/Carmel/Her2/Batch_4/HER2_4'),
        Path('/mnt/gipmed_new/Data/Breast/Carmel/Her2/Batch_5/HER2_5'),
        Path('/mnt/gipmed_new/Data/Breast/Carmel/Her2/Batch_6/HER2_6'),
    ]

    # slide_mapping = SlidesMapping.from_excel(path=Path('./scripts/output.xlsx'))
    slide_mapping = SlidesMapping.from_file_paths(paths=base_paths)
    slide_mapping.save_dataframe(path=Path('./output.xlsx'))
    slide_mapping.save_thumbnails(input_path=input_path, output_path=output_path, size=(800, 800))
